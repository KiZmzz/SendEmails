import json
import re
import smtplib
import sys
import threading
import time
import urllib.error
import urllib.parse
import urllib.request
from email.header import Header
from email.mime.text import MIMEText
from email.utils import formataddr, make_msgid, formatdate
import sys
from pathlib import Path

def get_resource_path(relative_path):
    if hasattr(sys, '_MEIPASS'):
        return Path(sys._MEIPASS) / relative_path
    return Path(__file__).parent / relative_path

import openpyxl
import keyring
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QUrl, QTimer, QSize
from PyQt5.QtGui import QIcon, QDesktopServices, QFont, QTextCharFormat, QColor
from PyQt5.QtWebEngineWidgets import QWebEngineView
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QFileDialog, QStackedWidget, QColorDialog
)

from qfluentwidgets import (
    FluentWindow, NavigationItemPosition, ScrollArea, LineEdit, 
    TextEdit, PushButton, PrimaryPushButton, ComboBox, PasswordLineEdit, 
    MessageBox, ListWidget, ProgressBar, StrongBodyLabel, InfoBar, InfoBarPosition,
    BodyLabel, CardWidget, SubtitleLabel, setTheme, Theme, FluentIcon as FIF,
    setThemeColor, PrimaryToolButton, TransparentToolButton, InfoBadge, InfoLevel,
    TitleLabel
)


EMAIL_PATTERN = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")

EMAIL_PROVIDERS = {
    "Outlook / Microsoft 365": {
        "mode": "graph",
        "description": "使用 Microsoft Graph OAuth 2.0 登录，适合 Outlook 和 Microsoft 365。",
    },
    "阿里企业邮箱": {
        "mode": "smtp",
        "server": "smtp.qiye.aliyun.com",
        "port": "465",
        "description": "使用 SMTP 登录，通常需要三方客户端安全密码。",
    },
    "QQ邮箱": {
        "mode": "smtp",
        "server": "smtp.qq.com",
        "port": "465",
        "description": "使用 SMTP 登录，通常需要授权码。",
    },
    "网易邮箱": {
        "mode": "smtp",
        "server": "smtp.163.com",
        "port": "465",
        "description": "使用 SMTP 登录，通常需要授权码。",
    },
    "Gmail": {
        "mode": "smtp",
        "server": "smtp.gmail.com",
        "port": "587",
        "description": "使用 SMTP 登录，通常需要应用专用密码。",
    },
    "其他 SMTP": {
        "mode": "smtp",
        "server": "",
        "port": "",
        "description": "手动填写 SMTP 服务器和端口。",
    },
}

EMAIL_HEADER_KEYWORDS = (
    "email", "e-mail", "mail", "邮箱", "邮箱地址", "电子邮箱", "收件邮箱", "收件人邮箱"
)

GRAPH_SCOPES = "offline_access openid profile User.Read Mail.Send"


class DeviceFlowCancelled(Exception):
    pass


class ConfigManager:
    APP_NAME = "SendEmailsTool"
    CONFIG_FILE = get_resource_path("accounts.json")
    
    @classmethod
    def load_accounts(cls):
        if not cls.CONFIG_FILE.exists():
            return {}
        try:
            with open(cls.CONFIG_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
            
    @classmethod
    def save_account(cls, config):
        account = config.get("account")
        if not account:
            return
            
        accounts = cls.load_accounts()
        
        # Extract sensitive info
        secret = ""
        config_to_save = dict(config)
        
        if config["mode"] == "smtp":
            secret = config_to_save.pop("password", "")
        else:
            token_data = config_to_save.pop("token_data", None)
            if token_data:
                secret = json.dumps(token_data)
                
        accounts[account] = config_to_save
        
        # Save JSON
        with open(cls.CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(accounts, f, ensure_ascii=False, indent=4)
            
        # Save Secret to Keychain
        if secret:
            try:
                keyring.set_password(cls.APP_NAME, account, secret)
            except Exception as e:
                print("Failed to save to keyring:", e)
                
    @classmethod
    def get_secret(cls, account, mode):
        try:
            secret = keyring.get_password(cls.APP_NAME, account)
            if not secret:
                return None
            if mode == "smtp":
                return secret
            else:
                return json.loads(secret)
        except Exception as e:
            print("Failed to get from keyring:", e)
            return None
            
    @classmethod
    def delete_account(cls, account):
        accounts = cls.load_accounts()
        if account in accounts:
            del accounts[account]
            with open(cls.CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump(accounts, f, ensure_ascii=False, indent=4)
        
        try:
            keyring.delete_password(cls.APP_NAME, account)
        except Exception:
            pass


class TemplateManager:
    TEMPLATE_FILE = get_resource_path("templates.json")
    
    @classmethod
    def load_templates(cls):
        if not cls.TEMPLATE_FILE.exists():
            return {}
        try:
            with open(cls.TEMPLATE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
            
    @classmethod
    def save_template(cls, name, subject, html_content):
        templates = cls.load_templates()
        templates[name] = {
            "subject": subject,
            "content": html_content
        }
        with open(cls.TEMPLATE_FILE, "w", encoding="utf-8") as f:
            json.dump(templates, f, ensure_ascii=False, indent=4)
            
    @classmethod
    def delete_template(cls, name):
        templates = cls.load_templates()
        if name in templates:
            del templates[name]
            with open(cls.TEMPLATE_FILE, "w", encoding="utf-8") as f:
                json.dump(templates, f, ensure_ascii=False, indent=4)


# ===============================
# Workers
# ===============================
class SharedAPI:
    @staticmethod
    def post_form(url, data):
        body = urllib.parse.urlencode(data).encode("utf-8")
        request = urllib.request.Request(
            url,
            data=body,
            method="POST",
            headers={"Content-Type": "application/x-www-form-urlencoded"},
        )
        return SharedAPI.perform_request(request)

    @staticmethod
    def graph_request(method, url, access_token, payload=None):
        headers = {"Authorization": f"Bearer {access_token}"}
        data = None
        if payload is not None:
            headers["Content-Type"] = "application/json"
            data = json.dumps(payload).encode("utf-8")
        request = urllib.request.Request(url, data=data, method=method, headers=headers)
        return SharedAPI.perform_request(request)

    @staticmethod
    def perform_request(request):
        try:
            with urllib.request.urlopen(request, timeout=30) as response:
                raw = response.read().decode("utf-8", errors="replace")
                return response.getcode(), json.loads(raw) if raw else {}
        except urllib.error.HTTPError as exc:
            raw = exc.read().decode("utf-8", errors="replace")
            try:
                payload = json.loads(raw) if raw else {}
            except json.JSONDecodeError:
                payload = {"error_description": raw or str(exc)}
            return exc.code, payload
        except Exception as exc:
            raise RuntimeError(str(exc)) from exc

    @staticmethod
    def format_oauth_error(payload):
        if not isinstance(payload, dict):
            return str(payload)
        error = payload.get("error", "oauth_error")
        description = payload.get("error_description") or payload.get("message") or "微软授权失败。"
        return f"{error}\n{description}"

    @staticmethod
    def format_graph_error(payload):
        if isinstance(payload, dict) and "error" in payload:
            error_block = payload["error"]
            if isinstance(error_block, dict):
                code = error_block.get("code", "graph_error")
                message = error_block.get("message", "Microsoft Graph 请求失败。")
                return f"{code}\n{message}"
        return str(payload)

    @staticmethod
    def create_smtp_client(smtp_server, smtp_port):
        if smtp_port == 465:
            client = smtplib.SMTP_SSL(smtp_server, smtp_port, timeout=30)
        else:
            client = smtplib.SMTP(smtp_server, smtp_port, timeout=30)
            client.ehlo()
            if smtp_port in (587, 25):
                client.starttls()
                client.ehlo()
        return client


class MicrosoftAuthWorker(QThread):
    flow_ready = pyqtSignal(dict)
    login_success = pyqtSignal(dict, dict)
    login_failed = pyqtSignal(str)
    
    def __init__(self, tenant, client_id):
        super().__init__()
        self.tenant = tenant
        self.client_id = client_id
        self.is_cancelled = False
        self.active_device_code = None

    def cancel(self):
        self.is_cancelled = True

    def run(self):
        try:
            flow = self.request_device_code()
            self.active_device_code = flow["device_code"]
            self.flow_ready.emit(flow)
            
            token_data = self.poll_device_code_token(flow)
            profile = self.fetch_graph_profile(token_data["access_token"])
            self.login_success.emit(token_data, profile)
        except DeviceFlowCancelled:
            pass # Cancelled silently
        except Exception as exc:
            self.login_failed.emit(str(exc))

    def request_device_code(self):
        url = f"https://login.microsoftonline.com/{self.tenant}/oauth2/v2.0/devicecode"
        status, payload = SharedAPI.post_form(url, {"client_id": self.client_id, "scope": GRAPH_SCOPES})
        if status >= 400:
            raise RuntimeError(SharedAPI.format_oauth_error(payload))
        return payload

    def poll_device_code_token(self, flow):
        url = f"https://login.microsoftonline.com/{self.tenant}/oauth2/v2.0/token"
        interval = int(flow.get("interval", 5))
        expires_at = time.time() + int(flow.get("expires_in", 900))

        while time.time() < expires_at:
            if self.is_cancelled:
                raise DeviceFlowCancelled()

            status, payload = SharedAPI.post_form(
                url,
                {
                    "grant_type": "urn:ietf:params:oauth:grant-type:device_code",
                    "client_id": self.client_id,
                    "device_code": flow["device_code"],
                },
            )

            if status < 400 and payload.get("access_token"):
                payload["obtained_at"] = time.time()
                return payload

            error = payload.get("error")
            if error == "authorization_pending":
                time.sleep(interval)
                continue
            if error == "slow_down":
                time.sleep(interval + 5)
                continue

            raise RuntimeError(SharedAPI.format_oauth_error(payload))

        raise RuntimeError("设备码已过期，请重新发起微软登录。")

    def fetch_graph_profile(self, access_token):
        status, payload = SharedAPI.graph_request(
            "GET",
            "https://graph.microsoft.com/v1.0/me?$select=displayName,mail,userPrincipalName",
            access_token,
        )
        if status >= 400:
            raise RuntimeError(SharedAPI.format_graph_error(payload))
        return payload


class SmtpTestWorker(QThread):
    success = pyqtSignal()
    failed = pyqtSignal(str)

    def __init__(self, config):
        super().__init__()
        self.config = config

    def run(self):
        try:
            with SharedAPI.create_smtp_client(self.config["smtp_server"], self.config["smtp_port"]) as client:
                client.login(self.config["account"], self.config["password"])
            self.success.emit()
        except Exception as exc:
            self.failed.emit(str(exc))


class EmailSendWorker(QThread):
    progress = pyqtSignal(int, int, int, int) # index, total, success, fail
    complete = pyqtSignal(bool, int, list) # connected, success_count, failed_list

    def __init__(self, config, subject, content, recipients, token_data=None, send_interval=0):
        super().__init__()
        self.config = config
        self.subject = subject
        self.content = content
        self.recipients = recipients
        self.token_data = token_data
        self.send_interval = send_interval
        self.mode = "graph" if token_data is not None else "smtp"

    def run(self):
        if self.mode == "graph":
            self.run_graph()
        else:
            self.run_smtp()

    def run_smtp(self):
        success_count = 0
        failed = []
        try:
            with SharedAPI.create_smtp_client(self.config["smtp_server"], self.config["smtp_port"]) as client:
                client.login(self.config["account"], self.config["password"])
                for index, recipient in enumerate(self.recipients, start=1):
                    try:
                        message = MIMEText(self.content, "html", "utf-8")
                        message["From"] = formataddr(
                            (str(Header(self.config["nickname"], "utf-8")), self.config["account"])
                        )
                        message["To"] = recipient
                        message["Subject"] = Header(self.subject, "utf-8")
                        message["Message-ID"] = make_msgid()
                        message["Date"] = formatdate()
                        client.sendmail(self.config["account"], [recipient], message.as_string())
                        success_count += 1
                    except Exception as exc:
                        failed.append((recipient, str(exc)))
                    self.progress.emit(index, len(self.recipients), success_count, len(failed))
                    if self.send_interval > 0 and index < len(self.recipients):
                        time.sleep(self.send_interval)
        except Exception as exc:
            self.complete.emit(False, 0, [(self.config["account"], str(exc))])
            return

        self.complete.emit(True, success_count, failed)

    def run_graph(self):
        success_count = 0
        failed = []
        try:
            self.refresh_graph_token_if_needed()
            token = self.token_data["access_token"]
        except Exception as exc:
            self.complete.emit(False, 0, [(self.config["account"], str(exc))])
            return

        for index, recipient in enumerate(self.recipients, start=1):
            try:
                payload = self.build_graph_message_payload(recipient, self.subject, self.content)
                status, response = SharedAPI.graph_request(
                    "POST",
                    "https://graph.microsoft.com/v1.0/me/sendMail",
                    token,
                    payload,
                )
                if status >= 400:
                    raise RuntimeError(SharedAPI.format_graph_error(response))
                success_count += 1
            except Exception as exc:
                failed.append((recipient, str(exc)))
            self.progress.emit(index, len(self.recipients), success_count, len(failed))

        self.complete.emit(True, success_count, failed)

    def build_graph_message_payload(self, recipient, subject, content):
        return {
            "message": {
                "subject": subject,
                "body": {
                    "contentType": "HTML",
                    "content": content,
                },
                "toRecipients": [
                    {
                        "emailAddress": {
                            "address": recipient,
                        }
                    }
                ],
            },
            "saveToSentItems": True,
        }

    def refresh_graph_token_if_needed(self):
        expires_in = int(self.token_data.get("expires_in", 3600))
        obtained_at = float(self.token_data.get("obtained_at", 0))
        if time.time() < obtained_at + expires_in - 120:
            return

        refresh_token = self.token_data.get("refresh_token")
        if not refresh_token:
            raise RuntimeError("微软授权已过期，请重新登录。")

        url = f"https://login.microsoftonline.com/{self.config['tenant']}/oauth2/v2.0/token"
        status, payload = SharedAPI.post_form(
            url,
            {
                "client_id": self.config['client_id'],
                "grant_type": "refresh_token",
                "refresh_token": refresh_token,
                "scope": GRAPH_SCOPES,
            },
        )
        if status >= 400 or "access_token" not in payload:
            raise RuntimeError(SharedAPI.format_oauth_error(payload))

        payload["obtained_at"] = time.time()
        if "refresh_token" not in payload:
            payload["refresh_token"] = refresh_token
        self.token_data = payload


# ===============================
# UI Components
# ===============================
class WidgetBase(ScrollArea):
    def __init__(self, text: str, object_name: str, parent=None):
        super().__init__(parent=parent)
        self.setObjectName(object_name)
        self.view = QWidget(self)
        self.vBoxLayout = QVBoxLayout(self.view)
        
        self.titleLabel = TitleLabel(text, self.view)
        self.vBoxLayout.addWidget(self.titleLabel)
        self.vBoxLayout.setContentsMargins(36, 36, 36, 36)
        self.vBoxLayout.setSpacing(18)
        self.vBoxLayout.setAlignment(Qt.AlignTop)
        
        self.setWidget(self.view)
        self.setWidgetResizable(True)
        self.setStyleSheet("QScrollArea {background-color: transparent; border: none;}")
        self.view.setStyleSheet("QWidget {background-color: transparent;}")


class AccountInterface(WidgetBase):
    authStatusChanged = pyqtSignal(bool, str) # is_graph_authed, display_text

    def __init__(self, parent=None):
        super().__init__("发件账号", "accountInterface", parent)
        self.token_data = None
        self.oauth_profile = None

        self.setup_ui()
        self.on_provider_changed(self.providerCombo.currentText())

    def setup_ui(self):
        # 0. 已保存账号
        card0 = CardWidget(self.view)
        layout0 = QVBoxLayout(card0)
        
        saved_layout = QHBoxLayout()
        saved_layout.addWidget(StrongBodyLabel("已保存账号", card0))
        self.savedCombo = ComboBox(card0)
        self.savedCombo.addItem("--- 请选择或新建 ---")
        self.savedCombo.setFixedWidth(220)
        
        self.saved_accounts_data = ConfigManager.load_accounts()
        self.savedCombo.addItems(list(self.saved_accounts_data.keys()))
        self.savedCombo.currentTextChanged.connect(self.on_saved_account_selected)
        saved_layout.addWidget(self.savedCombo)
        
        self.saveBtn = PushButton(FIF.SAVE, "保存当前配置", card0)
        self.saveBtn.clicked.connect(self.save_current_config)
        saved_layout.addWidget(self.saveBtn)
        
        self.delBtn = PushButton(FIF.DELETE, "删除", card0)
        self.delBtn.clicked.connect(self.delete_account)
        saved_layout.addWidget(self.delBtn)
        
        layout0.addLayout(saved_layout)
        self.vBoxLayout.addWidget(card0)

        # 1. 邮箱类型选择
        card1 = CardWidget(self.view)
        layout1 = QVBoxLayout(card1)
        
        type_layout = QHBoxLayout()
        type_layout.addWidget(StrongBodyLabel("邮箱类型", card1))
        self.providerCombo = ComboBox(card1)
        self.providerCombo.addItems(list(EMAIL_PROVIDERS.keys()))
        self.providerCombo.currentTextChanged.connect(self.on_provider_changed)
        type_layout.addWidget(self.providerCombo)
        layout1.addLayout(type_layout)
        
        self.providerDesc = BodyLabel("", card1)
        self.providerDesc.setWordWrap(True)
        layout1.addWidget(self.providerDesc)
        self.vBoxLayout.addWidget(card1)

        # 2. 基础账号信息
        card2 = CardWidget(self.view)
        layout2 = QVBoxLayout(card2)
        
        layout2.addWidget(StrongBodyLabel("发件邮箱账号", card2))
        self.accountEdit = LineEdit(card2)
        self.accountEdit.setPlaceholderText("example@domain.com")
        layout2.addWidget(self.accountEdit)

        layout2.addWidget(StrongBodyLabel("显示昵称（可选）", card2))
        self.nicknameEdit = LineEdit(card2)
        self.nicknameEdit.setPlaceholderText("留空则默认显示账号")
        layout2.addWidget(self.nicknameEdit)
        self.vBoxLayout.addWidget(card2)

        # 3. SMTP 设置区
        self.smtpCard = CardWidget(self.view)
        smtpLayout = QVBoxLayout(self.smtpCard)
        
        smtpLayout.addWidget(StrongBodyLabel("SMTP 服务器", self.smtpCard))
        self.smtpServerEdit = LineEdit(self.smtpCard)
        smtpLayout.addWidget(self.smtpServerEdit)
        
        smtpLayout.addWidget(StrongBodyLabel("SMTP 端口", self.smtpCard))
        self.smtpPortEdit = LineEdit(self.smtpCard)
        smtpLayout.addWidget(self.smtpPortEdit)
        
        smtpLayout.addWidget(StrongBodyLabel("密码 / 授权码", self.smtpCard))
        self.passwordEdit = PasswordLineEdit(self.smtpCard)
        self.passwordEdit.setPlaceholderText("请输入邮箱密码或三方授权码")
        smtpLayout.addWidget(self.passwordEdit)
        
        self.smtpLoginBtn = PushButton("测试 SMTP 登录", self.smtpCard)
        self.smtpLoginBtn.clicked.connect(self.test_smtp_login)
        smtpLayout.addWidget(self.smtpLoginBtn)
        self.vBoxLayout.addWidget(self.smtpCard)

        # 4. Microsoft Graph 设置区
        self.graphCard = CardWidget(self.view)
        graphLayout = QVBoxLayout(self.graphCard)
        
        graphLayout.addWidget(StrongBodyLabel("Microsoft App Client ID", self.graphCard))
        self.clientIdEdit = LineEdit(self.graphCard)
        graphLayout.addWidget(self.clientIdEdit)
        
        graphLayout.addWidget(StrongBodyLabel("租户类型", self.graphCard))
        self.tenantCombo = ComboBox(self.graphCard)
        self.tenantCombo.addItems(["common", "organizations", "consumers"])
        graphLayout.addWidget(self.tenantCombo)
        
        btnLayout = QHBoxLayout()
        self.oauthLoginBtn = PrimaryPushButton("微软设备码登录", self.graphCard)
        self.oauthLoginBtn.clicked.connect(self.start_microsoft_login)
        self.clearOauthBtn = PushButton("清除授权", self.graphCard)
        self.clearOauthBtn.clicked.connect(self.clear_microsoft_token)
        btnLayout.addWidget(self.oauthLoginBtn)
        btnLayout.addWidget(self.clearOauthBtn)
        graphLayout.addLayout(btnLayout)

        # Graph login status
        self.graphStatusBadge = InfoBadge.info("未授权", self.graphCard)
        graphLayout.addWidget(self.graphStatusBadge, alignment=Qt.AlignLeft)
        
        self.vBoxLayout.addWidget(self.graphCard)

        self.vBoxLayout.addStretch(1)

    def on_saved_account_selected(self, account_email):
        if account_email == "--- 请选择或新建 ---" or account_email not in self.saved_accounts_data:
            return
            
        config = self.saved_accounts_data[account_email]
        
        # Set Provider
        provider = config.get("provider")
        if provider:
            self.providerCombo.setCurrentText(provider)
            
        self.accountEdit.setText(config.get("account", ""))
        self.nicknameEdit.setText(config.get("nickname", ""))
        
        mode = config.get("mode")
        secret = ConfigManager.get_secret(account_email, mode)
        
        if mode == "smtp":
            self.smtpServerEdit.setText(config.get("smtp_server", ""))
            self.smtpPortEdit.setText(str(config.get("smtp_port", "")))
            if secret:
                self.passwordEdit.setText(secret)
            else:
                self.passwordEdit.clear()
        else:
            self.clientIdEdit.setText(config.get("client_id", ""))
            self.tenantCombo.setCurrentText(config.get("tenant", "common"))
            if secret:
                self.token_data = secret
                self.graphStatusBadge.setText("从系统凭据恢复授权")
                self.graphStatusBadge.setLevel(InfoLevel.SUCCESS)
                self.authStatusChanged.emit(True, "微软授权已恢复")
            else:
                self.token_data = None
                self.graphStatusBadge.setText("未授权")
                self.graphStatusBadge.setLevel(InfoLevel.INFO)
                self.authStatusChanged.emit(False, "等待微软授权")

    def save_current_config(self):
        try:
            config = self.get_config()
        except ValueError as e:
            InfoBar.error("配置错误", str(e), parent=self.window())
            return
            
        ConfigManager.save_account(config)
        InfoBar.success("保存成功", "账号配置已保存，凭证已安全加密", parent=self.window())
        
        # Refresh combo
        self.saved_accounts_data = ConfigManager.load_accounts()
        current = config["account"]
        self.savedCombo.clear()
        self.savedCombo.addItem("--- 请选择或新建 ---")
        self.savedCombo.addItems(list(self.saved_accounts_data.keys()))
        self.savedCombo.setCurrentText(current)

    def delete_account(self):
        account = self.savedCombo.currentText()
        if account == "--- 请选择或新建 ---" or account not in self.saved_accounts_data:
            return
            
        ConfigManager.delete_account(account)
        InfoBar.success("删除成功", f"账号 '{account}' 已彻底删除", parent=self.window())
        
        # Refresh combo
        self.saved_accounts_data = ConfigManager.load_accounts()
        self.savedCombo.blockSignals(True)
        self.savedCombo.clear()
        self.savedCombo.addItem("--- 请选择或新建 ---")
        self.savedCombo.addItems(list(self.saved_accounts_data.keys()))
        self.savedCombo.setCurrentIndex(0)
        self.savedCombo.blockSignals(False)

    def on_provider_changed(self, text):
        provider = EMAIL_PROVIDERS[text]
        self.providerDesc.setText(provider["description"])
        mode = provider["mode"]
        
        if mode == "smtp":
            self.smtpCard.show()
            self.graphCard.hide()
            self.smtpServerEdit.setText(provider.get("server", ""))
            self.smtpPortEdit.setText(provider.get("port", ""))
            self.authStatusChanged.emit(False, "SMTP 模式")
        else:
            self.smtpCard.hide()
            self.graphCard.show()
            self.authStatusChanged.emit(False if not self.token_data else True, "等待微软授权" if not self.token_data else "微软授权已完成")

    def test_smtp_login(self):
        account = self.accountEdit.text().strip()
        pwd = self.passwordEdit.text().strip()
        server = self.smtpServerEdit.text().strip()
        port = self.smtpPortEdit.text().strip()

        if not account or not pwd or not server or not port:
            InfoBar.error("配置错误", "请填写完整的 SMTP 登录信息", parent=self.window())
            return
            
        config = {
            "account": account,
            "password": pwd,
            "smtp_server": server,
            "smtp_port": int(port)
        }
        
        self.smtpLoginBtn.setEnabled(False)
        self.smtpLoginBtn.setText("验证中...")
        
        self.smtpWorker = SmtpTestWorker(config)
        self.smtpWorker.success.connect(self.on_smtp_success)
        self.smtpWorker.failed.connect(self.on_smtp_failed)
        self.smtpWorker.start()

    def on_smtp_success(self):
        self.smtpLoginBtn.setEnabled(True)
        self.smtpLoginBtn.setText("测试 SMTP 登录")
        InfoBar.success("登录成功", "SMTP 账号验证成功", parent=self.window())

    def on_smtp_failed(self, error):
        self.smtpLoginBtn.setEnabled(True)
        self.smtpLoginBtn.setText("测试 SMTP 登录")
        MessageBox("登录失败", f"SMTP 验证失败：\n{error}", self.window()).exec()

    def start_microsoft_login(self):
        client_id = self.clientIdEdit.text().strip()
        tenant = self.tenantCombo.currentText().strip()
        if not client_id:
            InfoBar.error("配置错误", "请输入 Microsoft App Client ID", parent=self.window())
            return

        self.oauthLoginBtn.setEnabled(False)
        self.oauthLoginBtn.setText("请求设备码...")
        self.graphStatusBadge.setLevel(InfoLevel.INFO)
        
        self.authWorker = MicrosoftAuthWorker(tenant, client_id)
        self.authWorker.flow_ready.connect(self.on_flow_ready)
        self.authWorker.login_success.connect(self.on_oauth_success)
        self.authWorker.login_failed.connect(self.on_oauth_failed)
        self.authWorker.start()

    def on_flow_ready(self, flow):
        self.oauthLoginBtn.setText("等待授权完成...")
        QDesktopServices.openUrl(QUrl(flow.get("verification_uri_complete", flow.get("verification_uri"))))
        
        msg = MessageBox("微软设备码登录", 
                         f"浏览器已自动打开登录页。\n请输入以下设备码并完成授权：\n\n{flow['user_code']}\n\n完成后本对话框将自动关闭。", 
                         self.window())
        msg.cancelButton.setText("取消授权")
        msg.yesButton.setText("复制设备码")
        msg.yesButton.clicked.connect(lambda: QApplication.clipboard().setText(flow['user_code']))
        msg.cancelButton.clicked.connect(self.cancel_auth)
        self.authMsgBox = msg
        msg.show()

    def cancel_auth(self):
        if hasattr(self, 'authWorker') and self.authWorker.isRunning():
            self.authWorker.cancel()
        self.oauthLoginBtn.setEnabled(True)
        self.oauthLoginBtn.setText("微软设备码登录")

    def on_oauth_success(self, token_data, profile):
        if hasattr(self, 'authMsgBox'):
            self.authMsgBox.accept()
            
        self.token_data = token_data
        self.oauth_profile = profile
        self.oauthLoginBtn.setEnabled(True)
        self.oauthLoginBtn.setText("微软设备码登录")
        
        sender = profile.get("mail") or profile.get("userPrincipalName")
        if sender:
            self.accountEdit.setText(sender)
        if profile.get("displayName") and not self.nicknameEdit.text().strip():
            self.nicknameEdit.setText(profile.get("displayName"))
            
        self.graphStatusBadge.setText(f"已授权: {sender}")
        self.graphStatusBadge.setLevel(InfoLevel.SUCCESS)
        self.authStatusChanged.emit(True, "微软授权已完成")
        InfoBar.success("授权成功", f"成功登录为: {sender}", parent=self.window())

    def on_oauth_failed(self, error):
        if hasattr(self, 'authMsgBox'):
            self.authMsgBox.reject()
        self.oauthLoginBtn.setEnabled(True)
        self.oauthLoginBtn.setText("微软设备码登录")
        self.graphStatusBadge.setLevel(InfoLevel.ERROR)
        self.graphStatusBadge.setText("授权失败")
        MessageBox("微软登录失败", error, self.window()).exec()

    def clear_microsoft_token(self):
        self.token_data = None
        self.oauth_profile = None
        self.graphStatusBadge.setText("未授权")
        self.graphStatusBadge.setLevel(InfoLevel.INFO)
        self.authStatusChanged.emit(False, "等待微软授权")
        InfoBar.info("已清除", "微软授权信息已清除", parent=self.window())

    def get_config(self):
        provider = EMAIL_PROVIDERS[self.providerCombo.currentText()]
        mode = provider["mode"]
        account = self.accountEdit.text().strip()
        nickname = self.nicknameEdit.text().strip() or account
        
        if not account:
            raise ValueError("发件账号不能为空")
            
        if mode == "smtp":
            pwd = self.passwordEdit.text().strip()
            server = self.smtpServerEdit.text().strip()
            port = self.smtpPortEdit.text().strip()
            if not pwd or not server or not port.isdigit():
                raise ValueError("SMTP 配置不完整或端口错误")
            return {
                "provider": self.providerCombo.currentText(),
                "mode": "smtp",
                "account": account,
                "nickname": nickname,
                "password": pwd,
                "smtp_server": server,
                "smtp_port": int(port)
            }
        else:
            client_id = self.clientIdEdit.text().strip()
            tenant = self.tenantCombo.currentText().strip()
            if not client_id:
                raise ValueError("Microsoft Client ID 不能为空")
            if not self.token_data:
                raise ValueError("请先点击'微软设备码登录'完成授权")
            
            return {
                "provider": self.providerCombo.currentText(),
                "mode": "graph",
                "account": account,
                "nickname": nickname,
                "client_id": client_id,
                "tenant": tenant,
                "token_data": self.token_data
            }


class ContentInterface(WidgetBase):
    def __init__(self, parent=None):
        super().__init__("邮件内容", "contentInterface", parent)

        # 0. 模板管理区域
        tplCard = CardWidget(self.view)
        tplLayout = QVBoxLayout(tplCard)
        
        rowLayout = QHBoxLayout()
        rowLayout.addWidget(StrongBodyLabel("模板管理", tplCard))
        
        self.tplCombo = ComboBox(tplCard)
        self.tplCombo.addItem("--- 选择已存模板 ---")
        self.tplCombo.setFixedWidth(220)
        self.templates_data = TemplateManager.load_templates()
        self.tplCombo.addItems(list(self.templates_data.keys()))
        self.tplCombo.currentTextChanged.connect(self.on_template_selected)
        rowLayout.addWidget(self.tplCombo)
        
        self.tplNameEdit = LineEdit(tplCard)
        self.tplNameEdit.setPlaceholderText("给新模板起个名字...")
        rowLayout.addWidget(self.tplNameEdit)
        
        self.saveTplBtn = PushButton(FIF.SAVE, "存为模板", tplCard)
        self.saveTplBtn.clicked.connect(self.save_current_template)
        rowLayout.addWidget(self.saveTplBtn)
        
        self.delTplBtn = PushButton(FIF.DELETE, "删除", tplCard)
        self.delTplBtn.clicked.connect(self.delete_template)
        rowLayout.addWidget(self.delTplBtn)
        
        tplLayout.addLayout(rowLayout)
        self.vBoxLayout.addWidget(tplCard)

        # 1. 邮件编辑器
        card = CardWidget(self.view)
        layout = QVBoxLayout(card)

        layout.addWidget(StrongBodyLabel("邮件主题", card))
        self.subjectEdit = LineEdit(card)
        self.subjectEdit.setText("邮件测试")
        layout.addWidget(self.subjectEdit)

        layout.addWidget(StrongBodyLabel("邮件正文", card))
        
        self.webView = QWebEngineView(card)
        editor_path = get_resource_path("editor.html")
        self.webView.setUrl(QUrl.fromLocalFile(str(editor_path.absolute())))
        self.webView.setMinimumHeight(400)
        
        layout.addWidget(self.webView)

        self.vBoxLayout.addWidget(card)

    def on_template_selected(self, tpl_name):
        if tpl_name == "--- 选择已存模板 ---" or tpl_name not in self.templates_data:
            return
            
        tpl = self.templates_data[tpl_name]
        self.subjectEdit.setText(tpl.get("subject", ""))
        
        # 将 HTML 内容安全地传回给 JS，使用 json.dumps 避免引号转义问题
        html_content = tpl.get("content", "")
        js_code = f"setHtmlContent({json.dumps(html_content)});"
        self.webView.page().runJavaScript(js_code)
        
        self.tplNameEdit.setText(tpl_name)
        InfoBar.success("导入成功", f"模板 '{tpl_name}' 已成功加载", parent=self.window())

    def save_current_template(self):
        tpl_name = self.tplNameEdit.text().strip()
        if not tpl_name:
            InfoBar.error("错误", "请输入要保存的模板名称", parent=self.window())
            return
            
        subject = self.subjectEdit.text().strip()
        
        # 异步获取 HTML 并保存
        self.webView.page().runJavaScript("getHtmlContent();", lambda html: self._do_save_template(tpl_name, subject, html))

    def _do_save_template(self, tpl_name, subject, html):
        content = str(html).strip() if html else ""
        if not content or content == "<p><br></p>":
            InfoBar.error("错误", "模板内容不能为空", parent=self.window())
            return
            
        TemplateManager.save_template(tpl_name, subject, content)
        InfoBar.success("保存成功", f"模板 '{tpl_name}' 已保存", parent=self.window())
        
        # Refresh combo
        self.templates_data = TemplateManager.load_templates()
        current = tpl_name
        self.tplCombo.blockSignals(True)
        self.tplCombo.clear()
        self.tplCombo.addItem("--- 选择已存模板 ---")
        self.tplCombo.addItems(list(self.templates_data.keys()))
        self.tplCombo.setCurrentText(current)
        self.tplCombo.blockSignals(False)

    def delete_template(self):
        tpl_name = self.tplCombo.currentText()
        if tpl_name == "--- 选择已存模板 ---" or tpl_name not in self.templates_data:
            return
            
        TemplateManager.delete_template(tpl_name)
        InfoBar.success("删除成功", f"模板 '{tpl_name}' 已被删除", parent=self.window())
        
        # Refresh combo
        self.templates_data = TemplateManager.load_templates()
        self.tplCombo.blockSignals(True)
        self.tplCombo.clear()
        self.tplCombo.addItem("--- 选择已存模板 ---")
        self.tplCombo.addItems(list(self.templates_data.keys()))
        self.tplCombo.setCurrentIndex(0)
        self.tplCombo.blockSignals(False)
        self.tplNameEdit.clear()


class SendInterface(WidgetBase):
    sendRequested = pyqtSignal()
    
    def __init__(self, parent=None):
        super().__init__("收件人与发送", "sendInterface", parent)
        self.recipients = []
        
        self.setup_ui()

    def setup_ui(self):
        # Tools
        card = CardWidget(self.view)
        layout = QVBoxLayout(card)
        
        # Add manual
        manualLayout = QHBoxLayout()
        self.manualEdit = LineEdit(card)
        self.manualEdit.setPlaceholderText("手动输入邮箱（支持批量文本自动提取）")
        self.addBtn = PushButton("提取并添加", card)
        self.addBtn.clicked.connect(self.add_manual)
        manualLayout.addWidget(self.manualEdit)
        manualLayout.addWidget(self.addBtn)
        layout.addLayout(manualLayout)
        
        # Actions
        actionLayout = QHBoxLayout()
        self.importExcelBtn = PushButton(FIF.DOCUMENT, "导入 Excel", card)
        self.importExcelBtn.clicked.connect(self.import_excel)
        self.clearBtn = PushButton(FIF.DELETE, "清空列表", card)
        self.clearBtn.clicked.connect(self.clear_recipients)
        actionLayout.addWidget(self.importExcelBtn)
        actionLayout.addWidget(self.clearBtn)
        actionLayout.addStretch(1)
        layout.addLayout(actionLayout)
        
        # List
        self.countLabel = StrongBodyLabel("当前收件人：0 个", card)
        layout.addWidget(self.countLabel)
        
        self.listWidget = ListWidget(card)
        self.listWidget.setMinimumHeight(300)
        layout.addWidget(self.listWidget)
        
        self.vBoxLayout.addWidget(card)

        # Send section
        sendCard = CardWidget(self.view)
        sendLayout = QVBoxLayout(sendCard)

        # Send interval
        intervalLayout = QHBoxLayout()
        intervalLayout.addWidget(StrongBodyLabel("发送间隔（秒）", sendCard))
        self.intervalEdit = LineEdit(sendCard)
        self.intervalEdit.setText("5")
        self.intervalEdit.setPlaceholderText("建议 3~10，防止被阿里限流")
        self.intervalEdit.setFixedWidth(180)
        intervalLayout.addWidget(self.intervalEdit)
        intervalLayout.addStretch(1)
        sendLayout.addLayout(intervalLayout)
        
        self.progressBar = ProgressBar(sendCard)
        self.progressBar.hide()
        sendLayout.addWidget(self.progressBar)
        
        self.statusLabel = BodyLabel("等待发送...", sendCard)
        sendLayout.addWidget(self.statusLabel)
        
        self.sendBtn = PrimaryPushButton(FIF.SEND, "一键发送所有邮件", sendCard)
        self.sendBtn.setMinimumHeight(40)
        font = self.sendBtn.font()
        font.setBold(True)
        font.setPointSize(12)
        self.sendBtn.setFont(font)
        self.sendBtn.clicked.connect(self.sendRequested.emit)
        sendLayout.addWidget(self.sendBtn)
        
        self.vBoxLayout.addWidget(sendCard)

    def get_send_interval(self):
        try:
            return max(0, float(self.intervalEdit.text().strip()))
        except ValueError:
            return 0

    def add_manual(self):
        text = self.manualEdit.text().strip()
        emails = self.extract_emails(text)
        if not emails:
            InfoBar.warning("未找到邮箱", "输入的内容中未包含有效邮箱地址", parent=self.window())
            return
        self.merge_recipients(emails)
        self.manualEdit.clear()
        
    def import_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "选择 Excel", "", "Excel Files (*.xlsx *.xlsm)")
        if not path:
            return
        try:
            emails = self.load_emails_from_excel(path)
            if not emails:
                InfoBar.warning("未找到邮箱", "未能在该 Excel 中识别到邮箱地址", parent=self.window())
            else:
                added = self.merge_recipients(emails)
                InfoBar.success("导入成功", f"识别到 {len(emails)} 个，新增 {added} 个邮箱", parent=self.window())
        except Exception as e:
            MessageBox("导入失败", str(e), self.window()).exec()

    def clear_recipients(self):
        self.recipients.clear()
        self.listWidget.clear()
        self.update_count()

    def extract_emails(self, text):
        return self.unique_preserve_order(EMAIL_PATTERN.findall(text))
        
    def load_emails_from_excel(self, file_path):
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        emails = []
        try:
            for sheet in wb.worksheets:
                email_col = None
                header_row = None
                
                # Find header
                for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), 1):
                    for col_idx, val in enumerate(row):
                        text = str(val).strip().lower() if val is not None else ""
                        if any(kw in text for kw in EMAIL_HEADER_KEYWORDS):
                            email_col = col_idx
                            header_row = row_idx
                            break
                    if email_col is not None:
                        break
                        
                if email_col is not None:
                    for row in sheet.iter_rows(min_row=header_row+1, values_only=True):
                        if email_col < len(row):
                            emails.extend(self.extract_emails(str(row[email_col] if row[email_col] else "")))
                else:
                    for row in sheet.iter_rows(values_only=True):
                        for val in row:
                            emails.extend(self.extract_emails(str(val if val else "")))
        finally:
            wb.close()
        return self.unique_preserve_order(emails)

    def unique_preserve_order(self, items):
        seen = set()
        unique = []
        for it in items:
            it = it.strip()
            if it and it.lower() not in seen:
                seen.add(it.lower())
                unique.append(it)
        return unique

    def merge_recipients(self, emails):
        before = len(self.recipients)
        self.recipients = self.unique_preserve_order(self.recipients + emails)
        self.listWidget.clear()
        self.listWidget.addItems(self.recipients)
        self.update_count()
        return len(self.recipients) - before

    def update_count(self):
        self.countLabel.setText(f"当前收件人：{len(self.recipients)} 个")


class EmailSenderApp(FluentWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("批量邮件发送工具")
        self.resize(960, 700)
        self.setMinimumWidth(800)
        self.setMinimumHeight(600)
        
        # SubInterfaces
        self.accountInterface = AccountInterface(self)
        self.contentInterface = ContentInterface(self)
        self.sendInterface = SendInterface(self)
        
        self.initNavigation()
        
        # Connections
        self.sendInterface.sendRequested.connect(self.start_sending)

    def initNavigation(self):
        self.addSubInterface(self.accountInterface, FIF.PEOPLE, "账号设置")
        self.addSubInterface(self.contentInterface, FIF.EDIT, "邮件内容")
        self.addSubInterface(self.sendInterface, FIF.SEND, "收件人与发送")
        
        self.navigationInterface.setExpandWidth(220)

    def start_sending(self):
        if not self.sendInterface.recipients:
            InfoBar.error("错误", "请先添加收件人", parent=self)
            self.switchTo(self.sendInterface)
            return
            
        subject = self.contentInterface.subjectEdit.text().strip()
        
        # 异步获取网页里的 HTML 源码，成功后回调 _continue_sending
        self.contentInterface.webView.page().runJavaScript(
            "getHtmlContent();", 
            lambda html: self._continue_sending(subject, html)
        )
        
    def _continue_sending(self, subject, raw_html):
        content = str(raw_html).strip() if raw_html else ""
        if not content or content == "<p><br></p>":
            InfoBar.error("错误", "邮件正文不能为空", parent=self)
            self.switchTo(self.contentInterface)
            return

        try:
            config = self.accountInterface.get_config()
        except ValueError as e:
            InfoBar.error("配置错误", str(e), parent=self)
            self.switchTo(self.accountInterface)
            return
            
        self.sendInterface.sendBtn.setEnabled(False)
        self.sendInterface.progressBar.show()
        self.sendInterface.progressBar.setMaximum(len(self.sendInterface.recipients))
        self.sendInterface.progressBar.setValue(0)
        self.sendInterface.statusLabel.setText("正在发送...")
        
        send_interval = self.sendInterface.get_send_interval()
        token_data = config.get("token_data")
        self.worker = EmailSendWorker(config, subject, content, self.sendInterface.recipients, token_data, send_interval)
        self.worker.progress.connect(self.on_send_progress)
        self.worker.complete.connect(self.on_send_complete)
        self.worker.start()

    def on_send_progress(self, index, total, success, fail):
        self.sendInterface.progressBar.setValue(index)
        self.sendInterface.statusLabel.setText(f"发送进度 {index}/{total}，成功 {success}，失败 {fail}。")

    def on_send_complete(self, connected, success, failed):
        self.sendInterface.sendBtn.setEnabled(True)
        if not connected:
            error = failed[0][1] if failed else "未知错误"
            self.sendInterface.statusLabel.setText("发送失败！")
            MessageBox("发送失败", error, self).exec()
            return
            
        fail_count = len(failed)
        if fail_count == 0:
            self.sendInterface.statusLabel.setText("发送完成，全部成功！")
            InfoBar.success("发送完成", f"全部发送成功，共 {success} 封。", parent=self, duration=5000)
        else:
            self.sendInterface.statusLabel.setText(f"发送完成，成功 {success} 封，失败 {fail_count} 封。")
            detail = "\n".join([f"{em}: {rs}" for em, rs in failed[:5]])
            if fail_count > 5:
                detail += f"\n... 还有 {fail_count - 5} 个失败"
            MessageBox("部分发送失败", f"成功 {success} 封，失败 {fail_count} 封。\n\n失败详情：\n{detail}", self).exec()


if __name__ == '__main__':
    # Enable high DPI scaling
    try:
        QApplication.setHighDpiScaleFactorRoundingPolicy(
            Qt.HighDpiScaleFactorRoundingPolicy.PassThrough)
    except AttributeError:
        pass
    QApplication.setAttribute(Qt.AA_EnableHighDpiScaling)
    QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps)

    app = QApplication(sys.argv)
    
    # 设置主题颜色
    setTheme(Theme.LIGHT)
    
    w = EmailSenderApp()
    w.show()
    sys.exit(app.exec())

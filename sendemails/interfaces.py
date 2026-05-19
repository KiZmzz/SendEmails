import json
import openpyxl
from PyQt6.QtCore import Qt, QUrl, QTimer, pyqtSignal
from PyQt6.QtGui import QDesktopServices, QFont, QTextCharFormat, QColor
from PyQt6.QtWebEngineWidgets import QWebEngineView
from PyQt6.QtWidgets import (
    QVBoxLayout, QHBoxLayout, QFileDialog,
    QTableWidgetItem, QHeaderView, QAbstractItemView, QDialog, QLabel, QDialogButtonBox,
    QWidget, QApplication,
)
from qfluentwidgets import (
    ScrollArea, LineEdit, SearchLineEdit, TextEdit, PushButton, PrimaryPushButton,
    ComboBox, PasswordLineEdit, MessageBox, TableWidget, ProgressBar,
    StrongBodyLabel, InfoBar, InfoBarPosition, BodyLabel, CardWidget, SubtitleLabel,
    FluentIcon as FIF, PrimaryToolButton, TransparentToolButton, InfoBadge, InfoLevel,
)

from .constants import (
    EMAIL_PATTERN, NAME_EMAIL_PATTERN, EMAIL_PROVIDERS,
    EMAIL_HEADER_KEYWORDS, NAME_HEADER_KEYWORDS,
)
from .utils import get_resource_path, ConfigManager, TemplateManager
from .workers import SmtpTestWorker, MicrosoftAuthWorker


class WidgetBase(ScrollArea):
    def __init__(self, text: str, object_name: str, parent=None):
        super().__init__(parent=parent)
        self.setObjectName(object_name)
        self.view = QWidget(self)
        self.vBoxLayout = QVBoxLayout(self.view)
        
        self.titleLabel = SubtitleLabel(text, self.view)
        self.vBoxLayout.addWidget(self.titleLabel)
        self.vBoxLayout.setContentsMargins(36, 12, 36, 36)
        self.vBoxLayout.setSpacing(8)
        self.vBoxLayout.setAlignment(Qt.AlignmentFlag.AlignTop)
        
        self.setWidget(self.view)
        self.setWidgetResizable(True)
        self.setStyleSheet("QScrollArea {background-color: transparent; border: none;}")
        self.view.setStyleSheet("QWidget {background-color: transparent;}")


class AccountInterface(WidgetBase):
    authStatusChanged = pyqtSignal(bool, str)

    def __init__(self, parent=None):
        super().__init__("账号设置", "accountInterface", parent)
        self.token_data = None
        self.oauth_profile = None

        self.setup_ui()
        self.on_provider_changed(self.providerCombo.currentText())

    def setup_ui(self):
        card0 = CardWidget(self.view)
        layout0 = QVBoxLayout(card0)
        
        saved_layout = QHBoxLayout()
        saved_layout.addWidget(StrongBodyLabel("已保存账号", card0))
        self.savedCombo = ComboBox(card0)
        self.savedCombo.addItem("--- 请选择或新建 ---")
        self.savedCombo.setFixedWidth(220)
        
        self.saved_accounts_data = ConfigManager.load_accounts()
        self.savedCombo.addItems(list(self.saved_accounts_data.keys()))
        self.savedCombo.currentTextChanged.connect(self.on_saved_account_selected)
        saved_layout.addWidget(self.savedCombo)
        
        self.saveBtn = PushButton(FIF.SAVE, "保存当前账号配置", card0)
        self.saveBtn.clicked.connect(self.save_current_config)
        saved_layout.addWidget(self.saveBtn)
        
        self.delBtn = PushButton(FIF.DELETE, "删除", card0)
        self.delBtn.clicked.connect(self.delete_account)
        saved_layout.addWidget(self.delBtn)
        
        layout0.addLayout(saved_layout)
        self.vBoxLayout.addWidget(card0)

        card1 = CardWidget(self.view)
        layout1 = QVBoxLayout(card1)
        
        type_layout = QHBoxLayout()
        type_layout.addWidget(StrongBodyLabel("邮箱类型", card1))
        self.providerCombo = ComboBox(card1)
        self.providerCombo.addItems(list(EMAIL_PROVIDERS.keys()))
        self.providerCombo.currentTextChanged.connect(self.on_provider_changed)
        type_layout.addWidget(self.providerCombo)
        layout1.addLayout(type_layout)
        
        self.providerDesc = BodyLabel("", card1)
        self.providerDesc.setWordWrap(True)
        layout1.addWidget(self.providerDesc)
        self.vBoxLayout.addWidget(card1)

        card2 = CardWidget(self.view)
        layout2 = QVBoxLayout(card2)
        
        layout2.addWidget(StrongBodyLabel("发件邮箱账号", card2))
        self.accountEdit = LineEdit(card2)
        self.accountEdit.setPlaceholderText("example@domain.com")
        layout2.addWidget(self.accountEdit)

        layout2.addWidget(StrongBodyLabel("显示昵称（可选）", card2))
        self.nicknameEdit = LineEdit(card2)
        self.nicknameEdit.setPlaceholderText("留空则默认显示账号")
        layout2.addWidget(self.nicknameEdit)
        self.vBoxLayout.addWidget(card2)

        self.smtpCard = CardWidget(self.view)
        smtpLayout = QVBoxLayout(self.smtpCard)
        
        smtpLayout.addWidget(StrongBodyLabel("SMTP 服务器", self.smtpCard))
        self.smtpServerEdit = LineEdit(self.smtpCard)
        smtpLayout.addWidget(self.smtpServerEdit)
        
        smtpLayout.addWidget(StrongBodyLabel("SMTP 端口", self.smtpCard))
        self.smtpPortEdit = LineEdit(self.smtpCard)
        smtpLayout.addWidget(self.smtpPortEdit)
        
        smtpLayout.addWidget(StrongBodyLabel("密码 / 授权码", self.smtpCard))
        self.passwordEdit = PasswordLineEdit(self.smtpCard)
        self.passwordEdit.setPlaceholderText("请输入邮箱密码或三方授权码")
        smtpLayout.addWidget(self.passwordEdit)
        
        self.smtpLoginBtn = PushButton("测试 SMTP 登录", self.smtpCard)
        self.smtpLoginBtn.clicked.connect(self.test_smtp_login)
        smtpLayout.addWidget(self.smtpLoginBtn)
        self.vBoxLayout.addWidget(self.smtpCard)

        self.graphCard = CardWidget(self.view)
        graphLayout = QVBoxLayout(self.graphCard)
        
        graphLayout.addWidget(StrongBodyLabel("Microsoft App Client ID", self.graphCard))
        self.clientIdEdit = LineEdit(self.graphCard)
        graphLayout.addWidget(self.clientIdEdit)
        
        graphLayout.addWidget(StrongBodyLabel("租户类型", self.graphCard))
        self.tenantCombo = ComboBox(self.graphCard)
        self.tenantCombo.addItems(["common", "organizations", "consumers"])
        graphLayout.addWidget(self.tenantCombo)
        
        btnLayout = QHBoxLayout()
        self.oauthLoginBtn = PrimaryPushButton("微软设备码登录", self.graphCard)
        self.oauthLoginBtn.clicked.connect(self.start_microsoft_login)
        self.clearOauthBtn = PushButton("清除授权", self.graphCard)
        self.clearOauthBtn.clicked.connect(self.clear_microsoft_token)
        btnLayout.addWidget(self.oauthLoginBtn)
        btnLayout.addWidget(self.clearOauthBtn)
        graphLayout.addLayout(btnLayout)

        self.graphStatusBadge = InfoBadge.info("未授权", self.graphCard)
        graphLayout.addWidget(self.graphStatusBadge, alignment=Qt.AlignmentFlag.AlignLeft)
        
        self.vBoxLayout.addWidget(self.graphCard)
        self.vBoxLayout.addStretch(1)

    def on_saved_account_selected(self, account_email):
        if account_email == "--- 请选择或新建 ---" or account_email not in self.saved_accounts_data:
            return
            
        config = self.saved_accounts_data[account_email]
        provider = config.get("provider")
        if provider:
            self.providerCombo.setCurrentText(provider)
            
        self.accountEdit.setText(config.get("account", ""))
        self.nicknameEdit.setText(config.get("nickname", ""))
        
        mode = config.get("mode")
        secret = ConfigManager.get_secret(account_email, mode)
        
        if mode == "smtp":
            self.smtpServerEdit.setText(config.get("smtp_server", ""))
            self.smtpPortEdit.setText(str(config.get("smtp_port", "")))
            if secret:
                self.passwordEdit.setText(secret)
            else:
                self.passwordEdit.clear()
        else:
            self.clientIdEdit.setText(config.get("client_id", ""))
            self.tenantCombo.setCurrentText(config.get("tenant", "common"))
            if secret:
                self.token_data = secret
                self.graphStatusBadge.setText("从系统凭据恢复授权")
                self.graphStatusBadge.setLevel(InfoLevel.SUCCESS)
                self.authStatusChanged.emit(True, "微软授权已恢复")
            else:
                self.token_data = None
                self.graphStatusBadge.setText("未授权")
                self.graphStatusBadge.setLevel(InfoLevel.INFO)
                self.authStatusChanged.emit(False, "等待微软授权")

    def save_current_config(self):
        try:
            config = self.get_config()
        except ValueError as e:
            InfoBar.error("配置错误", str(e), parent=self.window())
            return
            
        ConfigManager.save_account(config)
        InfoBar.success("保存成功", "账号配置已保存，凭证已安全加密", parent=self.window())
        
        self.saved_accounts_data = ConfigManager.load_accounts()
        current = config["account"]
        self.savedCombo.clear()
        self.savedCombo.addItem("--- 请选择或新建 ---")
        self.savedCombo.addItems(list(self.saved_accounts_data.keys()))
        self.savedCombo.setCurrentText(current)

    def delete_account(self):
        account = self.savedCombo.currentText()
        if account == "--- 请选择或新建 ---" or account not in self.saved_accounts_data:
            return
            
        ConfigManager.delete_account(account)
        InfoBar.success("删除成功", f"账号 '{account}' 已彻底删除", parent=self.window())
        
        self.saved_accounts_data = ConfigManager.load_accounts()
        self.savedCombo.blockSignals(True)
        self.savedCombo.clear()
        self.savedCombo.addItem("--- 请选择或新建 ---")
        self.savedCombo.addItems(list(self.saved_accounts_data.keys()))
        self.savedCombo.setCurrentIndex(0)
        self.savedCombo.blockSignals(False)

    def on_provider_changed(self, text):
        provider = EMAIL_PROVIDERS[text]
        self.providerDesc.setText(provider["description"])
        mode = provider["mode"]
        
        if mode == "smtp":
            self.smtpCard.show()
            self.graphCard.hide()
            self.smtpServerEdit.setText(provider.get("server", ""))
            self.smtpPortEdit.setText(provider.get("port", ""))
            self.authStatusChanged.emit(False, "SMTP 模式")
        else:
            self.smtpCard.hide()
            self.graphCard.show()
            self.authStatusChanged.emit(False if not self.token_data else True, "等待微软授权" if not self.token_data else "微软授权已完成")

    def test_smtp_login(self):
        account = self.accountEdit.text().strip()
        pwd = self.passwordEdit.text().strip()
        server = self.smtpServerEdit.text().strip()
        port = self.smtpPortEdit.text().strip()

        if not account or not pwd or not server or not port:
            InfoBar.error("配置错误", "请填写完整的 SMTP 登录信息", parent=self.window())
            return
            
        config = {
            "account": account,
            "password": pwd,
            "smtp_server": server,
            "smtp_port": int(port)
        }
        
        self.smtpLoginBtn.setEnabled(False)
        self.smtpLoginBtn.setText("验证中...")
        
        self.smtpWorker = SmtpTestWorker(config)
        self.smtpWorker.success.connect(self.on_smtp_success)
        self.smtpWorker.failed.connect(self.on_smtp_failed)
        self.smtpWorker.start()

    def on_smtp_success(self):
        self.smtpLoginBtn.setEnabled(True)
        self.smtpLoginBtn.setText("测试 SMTP 登录")
        InfoBar.success("登录成功", "SMTP 账号验证成功", parent=self.window())

    def on_smtp_failed(self, error):
        self.smtpLoginBtn.setEnabled(True)
        self.smtpLoginBtn.setText("测试 SMTP 登录")
        MessageBox("登录失败", f"SMTP 验证失败：\n{error}", self.window()).exec()

    def start_microsoft_login(self):
        client_id = self.clientIdEdit.text().strip()
        tenant = self.tenantCombo.currentText().strip()
        if not client_id:
            InfoBar.error("配置错误", "请输入 Microsoft App Client ID", parent=self.window())
            return

        self.oauthLoginBtn.setEnabled(False)
        self.oauthLoginBtn.setText("请求设备码...")
        self.graphStatusBadge.setLevel(InfoLevel.INFO)
        
        self.authWorker = MicrosoftAuthWorker(tenant, client_id)
        self.authWorker.flow_ready.connect(self.on_flow_ready)
        self.authWorker.login_success.connect(self.on_oauth_success)
        self.authWorker.login_failed.connect(self.on_oauth_failed)
        self.authWorker.start()

    def on_flow_ready(self, flow):
        self.oauthLoginBtn.setText("等待授权完成...")
        QDesktopServices.openUrl(QUrl(flow.get("verification_uri_complete", flow.get("verification_uri"))))
        
        msg = MessageBox("微软设备码登录", 
                         f"浏览器已自动打开登录页。\n请输入以下设备码并完成授权：\n\n{flow['user_code']}\n\n完成后本对话框将自动关闭。", 
                         self.window())
        msg.cancelButton.setText("取消授权")
        msg.yesButton.setText("复制设备码")
        msg.yesButton.clicked.connect(lambda: QApplication.clipboard().setText(flow['user_code']))
        msg.cancelButton.clicked.connect(self.cancel_auth)
        self.authMsgBox = msg
        msg.show()

    def cancel_auth(self):
        if hasattr(self, 'authWorker') and self.authWorker.isRunning():
            self.authWorker.cancel()
        self.oauthLoginBtn.setEnabled(True)
        self.oauthLoginBtn.setText("微软设备码登录")

    def on_oauth_success(self, token_data, profile):
        if hasattr(self, 'authMsgBox'):
            self.authMsgBox.accept()
            
        self.token_data = token_data
        self.oauth_profile = profile
        self.oauthLoginBtn.setEnabled(True)
        self.oauthLoginBtn.setText("微软设备码登录")
        
        sender = profile.get("mail") or profile.get("userPrincipalName")
        if sender:
            self.accountEdit.setText(sender)
        if profile.get("displayName") and not self.nicknameEdit.text().strip():
            self.nicknameEdit.setText(profile.get("displayName"))
            
        self.graphStatusBadge.setText(f"已授权: {sender}")
        self.graphStatusBadge.setLevel(InfoLevel.SUCCESS)
        self.authStatusChanged.emit(True, "微软授权已完成")
        InfoBar.success("授权成功", f"成功登录为: {sender}", parent=self.window())

    def on_oauth_failed(self, error):
        if hasattr(self, 'authMsgBox'):
            self.authMsgBox.reject()
        self.oauthLoginBtn.setEnabled(True)
        self.oauthLoginBtn.setText("微软设备码登录")
        self.graphStatusBadge.setLevel(InfoLevel.ERROR)
        self.graphStatusBadge.setText("授权失败")
        MessageBox("微软登录失败", error, self.window()).exec()

    def clear_microsoft_token(self):
        self.token_data = None
        self.oauth_profile = None
        self.graphStatusBadge.setText("未授权")
        self.graphStatusBadge.setLevel(InfoLevel.INFO)
        self.authStatusChanged.emit(False, "等待微软授权")
        InfoBar.info("已清除", "微软授权信息已清除", parent=self.window())

    def get_config(self):
        provider = EMAIL_PROVIDERS[self.providerCombo.currentText()]
        mode = provider["mode"]
        account = self.accountEdit.text().strip()
        nickname = self.nicknameEdit.text().strip() or account
        
        if not account:
            raise ValueError("发件账号不能为空")
            
        if mode == "smtp":
            pwd = self.passwordEdit.text().strip()
            server = self.smtpServerEdit.text().strip()
            port = self.smtpPortEdit.text().strip()
            if not pwd or not server or not port.isdigit():
                raise ValueError("SMTP 配置不完整或端口错误")
            return {
                "provider": self.providerCombo.currentText(),
                "mode": "smtp",
                "account": account,
                "nickname": nickname,
                "password": pwd,
                "smtp_server": server,
                "smtp_port": int(port)
            }
        else:
            client_id = self.clientIdEdit.text().strip()
            tenant = self.tenantCombo.currentText().strip()
            if not client_id:
                raise ValueError("Microsoft Client ID 不能为空")
            if not self.token_data:
                raise ValueError("请先点击'微软设备码登录'完成授权")
            
            return {
                "provider": self.providerCombo.currentText(),
                "mode": "graph",
                "account": account,
                "nickname": nickname,
                "client_id": client_id,
                "tenant": tenant,
                "token_data": self.token_data
            }


class ContentInterface(WidgetBase):
    def __init__(self, parent=None):
        super().__init__("邮件内容", "contentInterface", parent)

        tplCard = CardWidget(self.view)
        tplLayout = QVBoxLayout(tplCard)
        
        rowLayout = QHBoxLayout()
        rowLayout.addWidget(StrongBodyLabel("模板管理", tplCard))
        
        self.tplCombo = ComboBox(tplCard)
        self.tplCombo.addItem("--- 选择已存模板 ---")
        self.tplCombo.setFixedWidth(220)
        self.templates_data = TemplateManager.load_templates()
        self.tplCombo.addItems(list(self.templates_data.keys()))
        self.tplCombo.currentTextChanged.connect(self.on_template_selected)
        rowLayout.addWidget(self.tplCombo)
        
        self.tplNameEdit = LineEdit(tplCard)
        self.tplNameEdit.setPlaceholderText("给新模板起个名字...")
        rowLayout.addWidget(self.tplNameEdit)
        
        self.saveTplBtn = PushButton(FIF.SAVE, "存为模板", tplCard)
        self.saveTplBtn.clicked.connect(self.save_current_template)
        rowLayout.addWidget(self.saveTplBtn)
        
        self.delTplBtn = PushButton(FIF.DELETE, "删除", tplCard)
        self.delTplBtn.clicked.connect(self.delete_template)
        rowLayout.addWidget(self.delTplBtn)
        
        tplLayout.addLayout(rowLayout)
        self.vBoxLayout.addWidget(tplCard)

        card = CardWidget(self.view)
        layout = QVBoxLayout(card)

        layout.addWidget(StrongBodyLabel("邮件主题", card))
        self.subjectEdit = LineEdit(card)
        self.subjectEdit.setText("")
        layout.addWidget(self.subjectEdit)

        layout.addWidget(StrongBodyLabel("邮件正文", card))
        
        self.webView = QWebEngineView(card)
        editor_path = get_resource_path("editor.html")
        self.webView.setUrl(QUrl.fromLocalFile(str(editor_path.absolute())))
        self.webView.setMinimumHeight(400)
        
        layout.addWidget(self.webView)
        self.vBoxLayout.addWidget(card)

    def on_template_selected(self, tpl_name):
        if tpl_name == "--- 选择已存模板 ---" or tpl_name not in self.templates_data:
            return
            
        tpl = self.templates_data[tpl_name]
        self.subjectEdit.setText(tpl.get("subject", ""))
        html_content = tpl.get("content", "")
        js_code = f"setHtmlContent({json.dumps(html_content)});"
        self.webView.page().runJavaScript(js_code)
        
        self.tplNameEdit.setText(tpl_name)
        InfoBar.success("导入成功", f"模板 '{tpl_name}' 已成功加载", parent=self.window())

    def save_current_template(self):
        tpl_name = self.tplNameEdit.text().strip()
        if not tpl_name:
            InfoBar.error("错误", "请输入要保存的模板名称", parent=self.window())
            return
            
        subject = self.subjectEdit.text().strip()
        self.webView.page().runJavaScript("getHtmlContent();", lambda html: self._do_save_template(tpl_name, subject, html))

    def _do_save_template(self, tpl_name, subject, html):
        content = str(html).strip() if html else ""
        if not content or content == "<p><br></p>":
            InfoBar.error("错误", "模板内容不能为空", parent=self.window())
            return
            
        TemplateManager.save_template(tpl_name, subject, content)
        InfoBar.success("保存成功", f"模板 '{tpl_name}' 已保存", parent=self.window())
        
        self.templates_data = TemplateManager.load_templates()
        current = tpl_name
        self.tplCombo.blockSignals(True)
        self.tplCombo.clear()
        self.tplCombo.addItem("--- 选择已存模板 ---")
        self.tplCombo.addItems(list(self.templates_data.keys()))
        self.tplCombo.setCurrentText(current)
        self.tplCombo.blockSignals(False)

    def delete_template(self):
        tpl_name = self.tplCombo.currentText()
        if tpl_name == "--- 选择已存模板 ---" or tpl_name not in self.templates_data:
            return
            
        TemplateManager.delete_template(tpl_name)
        InfoBar.success("删除成功", f"模板 '{tpl_name}' 已被删除", parent=self.window())
        
        self.templates_data = TemplateManager.load_templates()
        self.tplCombo.blockSignals(True)
        self.tplCombo.clear()
        self.tplCombo.addItem("--- 选择已存模板 ---")
        self.tplCombo.addItems(list(self.templates_data.keys()))
        self.tplCombo.setCurrentIndex(0)
        self.tplCombo.blockSignals(False)
        self.tplNameEdit.clear()


class SendInterface(WidgetBase):
    sendRequested = pyqtSignal()
    
    def __init__(self, parent=None):
        super().__init__("收件人与发送", "sendInterface", parent)
        self.recipients = []
        self.setup_ui()

    def setup_ui(self):
        card = CardWidget(self.view)
        layout = QVBoxLayout(card)
        
        manualLayout = QHBoxLayout()
        self.manualEdit = LineEdit(card)
        self.manualEdit.setPlaceholderText("手动输入邮箱，例如：张三zhangsan@example.com")
        self.addBtn = PushButton("提取并添加", card)
        self.addBtn.clicked.connect(self.add_manual)
        manualLayout.addWidget(self.manualEdit)
        manualLayout.addWidget(self.addBtn)
        layout.addLayout(manualLayout)
        
        actionLayout = QHBoxLayout()
        self.importExcelBtn = PushButton(FIF.DOCUMENT, "导入 Excel", card)
        self.importExcelBtn.clicked.connect(self.import_excel)
        self.clearBtn = PushButton(FIF.DELETE, "清空列表", card)
        self.clearBtn.clicked.connect(self.clear_recipients)
        self.refreshBtn = PushButton(FIF.SYNC, "刷新", card)
        self.refreshBtn.clicked.connect(self._refresh_table)
        actionLayout.addWidget(self.importExcelBtn)
        actionLayout.addWidget(self.clearBtn)
        actionLayout.addWidget(self.refreshBtn)
        actionLayout.addStretch(1)
        self.searchEdit = SearchLineEdit(card)
        self.searchEdit.setPlaceholderText("搜索收件人...")
        self.searchEdit.setFixedWidth(240)
        self.searchEdit.textChanged.connect(self._filter_recipients)
        actionLayout.addWidget(self.searchEdit)
        layout.addLayout(actionLayout)
        
        self.countLabel = StrongBodyLabel("当前收件人：0 个", card)
        layout.addWidget(self.countLabel)
        
        self.table = TableWidget(card)
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(["收件人", "收件邮箱", "操作"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(0, 200)
        self.table.setColumnWidth(2, 100)
        self.table.setMinimumHeight(300)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.verticalHeader().hide()
        self.table.setAlternatingRowColors(True)
        self.table.setBorderVisible(True)
        self.table.verticalHeader().setDefaultSectionSize(36)
        self.table.setShowGrid(True)
        layout.addWidget(self.table)
        
        self.vBoxLayout.addWidget(card)

        sendCard = CardWidget(self.view)
        sendLayout = QVBoxLayout(sendCard)

        intervalLayout = QHBoxLayout()
        intervalLayout.addWidget(StrongBodyLabel("发送间隔（秒）", sendCard))
        self.intervalEdit = LineEdit(sendCard)
        self.intervalEdit.setText("5")
        self.intervalEdit.setPlaceholderText("建议 3~10，防止被阿里限流")
        self.intervalEdit.setFixedWidth(180)
        intervalLayout.addWidget(self.intervalEdit)
        intervalLayout.addStretch(1)
        sendLayout.addLayout(intervalLayout)
        
        self.progressBar = ProgressBar(sendCard)
        self.progressBar.hide()
        sendLayout.addWidget(self.progressBar)
        
        self.statusLabel = BodyLabel("等待发送...", sendCard)
        sendLayout.addWidget(self.statusLabel)
        
        self.sendBtn = PrimaryPushButton(FIF.SEND, "一键发送所有邮件", sendCard)
        self.sendBtn.setMinimumHeight(40)
        font = self.sendBtn.font()
        font.setBold(True)
        font.setPointSize(12)
        self.sendBtn.setFont(font)
        self.sendBtn.clicked.connect(self.sendRequested.emit)
        sendLayout.addWidget(self.sendBtn)
        
        self.vBoxLayout.addWidget(sendCard)

    def get_send_interval(self):
        try:
            return max(0, float(self.intervalEdit.text().strip()))
        except ValueError:
            return 0

    def add_manual(self):
        text = self.manualEdit.text().strip()
        pairs = NAME_EMAIL_PATTERN.findall(text)
        if not pairs:
            InfoBar.warning("未找到邮箱", "输入的内容中未包含有效邮箱地址", parent=self.window())
            return
        recipients = []
        for n, e in pairs:
            name = n.strip().rstrip(",").rstrip("，") or "未知"
            recipients.append((name, e))
        self.merge_recipients(recipients)
        self.manualEdit.clear()
        
    def import_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "选择 Excel", "", "Excel Files (*.xlsx *.xlsm)")
        if not path:
            return
        try:
            recipients = self.load_recipients_from_excel(path)
            if not recipients:
                InfoBar.warning("未找到邮箱", "未能在该 Excel 中识别到邮箱地址", parent=self.window())
            else:
                added = self.merge_recipients(recipients)
                InfoBar.success("导入成功", f"识别到 {len(recipients)} 个，新增 {added} 个", parent=self.window())
        except Exception as e:
            MessageBox("导入失败", str(e), self.window()).exec()

    def clear_recipients(self):
        self.recipients.clear()
        self.table.setRowCount(0)
        self.searchEdit.clear()
        self.update_count()

    def extract_emails(self, text):
        return self.unique_preserve_order(EMAIL_PATTERN.findall(text))
        
    def load_recipients_from_excel(self, file_path):
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        results = []
        try:
            for sheet in wb.worksheets:
                name_col = None
                email_col = None
                header_row = None
                
                for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), 1):
                    for col_idx, val in enumerate(row):
                        text = str(val).strip().lower() if val is not None else ""
                        if any(kw in text for kw in EMAIL_HEADER_KEYWORDS):
                            email_col = col_idx
                            header_row = row_idx
                        if any(kw in text for kw in NAME_HEADER_KEYWORDS):
                            name_col = col_idx
                            header_row = row_idx
                    if email_col is not None:
                        break
                
                if email_col is not None:
                    for row in sheet.iter_rows(min_row=header_row+1, values_only=True):
                        email_val = str(row[email_col]).strip() if email_col < len(row) and row[email_col] else ""
                        emails = EMAIL_PATTERN.findall(email_val)
                        if not emails:
                            continue
                        email = emails[0]
                        name = ""
                        if name_col is not None and name_col < len(row) and row[name_col]:
                            name = str(row[name_col]).strip()
                        if not name:
                            name = "未知"
                        results.append((name, email))
                else:
                    for row in sheet.iter_rows(values_only=True):
                        for val in row:
                            emails = EMAIL_PATTERN.findall(str(val if val else ""))
                            if emails:
                                results.append(("未知", emails[0]))
        finally:
            wb.close()
        return self.unique_preserve_order(results)

    def unique_preserve_order(self, items):
        seen = set()
        unique = []
        for it in items:
            if isinstance(it, str):
                it = it.strip()
                if it and it.lower() not in seen:
                    seen.add(it.lower())
                    unique.append(it)
            else:
                name, email = it
                email = email.strip()
                if email and email.lower() not in seen:
                    seen.add(email.lower())
                    unique.append((name.strip() or "未知", email))
        return unique

    def merge_recipients(self, new_recipients):
        before = len(self.recipients)
        all_items = self.recipients + new_recipients
        self.recipients = self.unique_preserve_order(all_items)
        self._populate_table()
        self.update_count()
        return len(self.recipients) - before

    def _filter_recipients(self, text):
        text = text.strip().lower()
        for row in range(self.table.rowCount()):
            nameItem = self.table.item(row, 0)
            emailItem = self.table.item(row, 1)
            if not text:
                self.table.setRowHidden(row, False)
            else:
                match = (text in nameItem.text().lower() or
                         text in emailItem.text().lower())
                self.table.setRowHidden(row, not match)

    def _refresh_table(self):
        self.searchEdit.clear()
        self._populate_table()
        self.update_count()

    def _populate_table(self):
        self.table.setRowCount(len(self.recipients))
        for row, (name, email) in enumerate(self.recipients):
            nameItem = QTableWidgetItem(name)
            nameItem.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row, 0, nameItem)
            emailItem = QTableWidgetItem(email)
            emailItem.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row, 1, emailItem)

            btnWidget = QWidget()
            btnLayout = QHBoxLayout(btnWidget)
            btnLayout.setContentsMargins(8, 4, 8, 4)
            btnLayout.setSpacing(6)

            editBtn = TransparentToolButton(FIF.EDIT, btnWidget)
            editBtn.setToolTip("修改")
            editBtn.clicked.connect(lambda checked, r=row: self._edit_recipient(r))
            delBtn = TransparentToolButton(FIF.DELETE, btnWidget)
            delBtn.setToolTip("删除")
            delBtn.clicked.connect(lambda checked, r=row: self._delete_recipient(r))

            btnLayout.addWidget(editBtn)
            btnLayout.addWidget(delBtn)
            btnLayout.addStretch()
            self.table.setCellWidget(row, 2, btnWidget)

    def _edit_recipient(self, row):
        name, email = self.recipients[row]
        dialog = QDialog(self.window())
        dialog.setWindowTitle("修改收件人")
        dialog.resize(360, 180)
        layout = QVBoxLayout(dialog)

        nameLabel = QLabel("收件人名称：")
        nameEdit = LineEdit(dialog)
        nameEdit.setText(name)
        layout.addWidget(nameLabel)
        layout.addWidget(nameEdit)

        emailLabel = QLabel("收件邮箱：")
        emailEdit = LineEdit(dialog)
        emailEdit.setText(email)
        layout.addWidget(emailLabel)
        layout.addWidget(emailEdit)

        btnBox = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel, dialog
        )
        btnBox.accepted.connect(dialog.accept)
        btnBox.rejected.connect(dialog.reject)
        layout.addWidget(btnBox)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            newName = nameEdit.text().strip() or "未知"
            newEmail = emailEdit.text().strip()
            if not EMAIL_PATTERN.fullmatch(newEmail):
                InfoBar.warning("邮箱无效", "请输入正确的邮箱地址", parent=self.window())
                return
            self.recipients[row] = (newName, newEmail)
            self._populate_table()
            self.update_count()

    def _delete_recipient(self, row):
        self.recipients.pop(row)
        self._populate_table()
        self.update_count()

    def update_count(self):
        self.countLabel.setText(f"当前收件人：{len(self.recipients)} 个")
